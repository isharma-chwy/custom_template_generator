import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Protection
from openpyxl.utils import column_index_from_string, quote_sheetname
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path


def create_excel_template(template_df: pd.DataFrame,
                          template_file_path: str = 'template.xlsx',
                          template_password: str = None,
                          output_path: str = './output/',
                          sheet_name: str = 'Sheet1',
                          editable_columns: list = [],
                          editable_cells: list = [],
                          start_cell: str = 'A2',
                          col_val_for_name: str = None,
                          cols_to_hide: list = [],
                          template_name: str = "template",
                          date_cell: str = 'D1'
                          )-> None:
    """Creates an excel template from a pandas dataframe using the template provided in template_file_path.
    Starts at the start_cell and fills the template's sheet_name with the data from the dataframe. (set to A2 and Sheet1 by default)
    Sets columns in the editable_columns list to be editable. (eg. A, B, C)
    Sets cells in the editable_cells list to be editable. (eg. A1, B2, C3)
    Hides columns in the cols_to_hide list. (eg. A, B, C)
    The template is saved in the output_path folder as (col_val_for_name)_(date_today)_(template_name).xlsx."""

    editable_columns = [col.upper() for col in editable_columns]
    cols_to_hide = [col.upper() for col in cols_to_hide]

    output_path = Path(output_path)

    if not output_path.exists():
        output_path.mkdir(parents=True, exist_ok=True)

    start_col, start_row = coordinate_from_string(start_cell)
    # print(start_row, start_col)
    start_col = column_index_from_string(start_col)
    num_rows = template_df.shape[0]
    num_cols = template_df.shape[1]

    #  Choose column to use for the name of the file and to split the data into multiple files
    if col_val_for_name is None:
        col_val_for_name = template_df[template_df.columns[0]].values[0]
    elif template_df.shape[0] > 1:
        col_val_for_name = template_df[col_val_for_name].values[0]
    else:
        raise ValueError("The dataframe is empty. Please check the query.")

    template_df_headers = list(template_df.columns)

    template_dict_list = template_df.to_dict('records')



    work_book = load_workbook(filename=template_file_path, read_only=False)
    wb_copy = work_book

    date_string = datetime.strftime(datetime.now(), '%d-%m-%Y')
    filename = f"{template_name}_{col_val_for_name}_{date_string}.xlsx"
    complete_file_name = os.path.join(output_path, filename)
    wb_copy.save(complete_file_name)
    work_book.close()

    work_book = load_workbook(filename=complete_file_name, read_only=False)

    work_sheet = work_book["Instructions"]
    work_sheet.protection.sheet = True

    work_sheet = work_book["ref"]
    work_sheet[date_cell].value = datetime.strftime(datetime.now(), '%m-%d-%Y')
    work_sheet.protection.sheet = True


    if template_password is not None:
        work_sheet.protection.password = template_password

    work_sheet = work_book[f"{sheet_name}"]
    work_sheet.protection.sheet = False

    # # write the headers to the excel sheet as per the Query - this is not needed if the headers are already in the template
    # for col_num in range(1,num_cols+1):
    #     work_sheet.cell(row=start_row-1, column=col_num).value = template_df_headers[col_num-1]
    #     work_sheet.cell(row=start_row-1, column=col_num).alignment = Alignment(horizontal="center", vertical="center" , wrap_text=True)

    # Write the data to the sheet under the headers
    for row_num in range(num_rows):
        for col_num in range(1,num_cols+1):
            work_sheet.cell(row=row_num + start_row, column=col_num).value = template_dict_list[row_num][template_df_headers[col_num-1]]
            work_sheet.cell(row=row_num + start_row, column=col_num).alignment = Alignment(horizontal="center")
    
    # Hide any columns in the cols_to_hide list
    for col in cols_to_hide:
        work_sheet.column_dimensions[col].hidden= True

    # unlock the cells in the editable_columns list and the cells in the editable_cells list
    for row_num in range(num_rows):
        for col in editable_columns:
            work_sheet.cell(row=row_num + start_row, column=column_index_from_string(col)).protection = Protection(locked=False)

    for cell in editable_cells:
        work_sheet[cell].protection = Protection(locked=False)


    # Add the drop down list to the cells in required columns
    # data_validation = DataValidation(type="list",formula1="{0}!$A$2:$A$3".format(quote_sheetname('ref')))
    # data_validation.add('J1:J5000')
    # data_validation.add('L1:L5000')

    work_sheet.protection.sheet = True
    work_sheet.protection.enable()
    work_book.save(complete_file_name)
    work_book.close()